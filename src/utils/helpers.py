import os
import sys

import httpx
import openpyxl
from openpyxl import Workbook

from src.utils.constants import HEADERS, MAIN_URL, NTLM_AUTH


async def get_response(link: str) -> httpx.Response:
    async with httpx.AsyncClient(
        http2=True,
        limits=httpx.Limits(
            max_connections=100,
            max_keepalive_connections=100,
        ),
    ) as client:
        response = await client.get(link, headers=HEADERS, auth=NTLM_AUTH)
        return response


def get_url_stop(link_up: str, date: str, shift: str) -> str:
    """Faster and more efficient version of get_url"""
    # return f"http://ots.app.pmi/db.aspx?table=sp_PeriodEquipmentData&act=query&eoa=x&db_Line=ID01-SE-CP-L0{link_up}&db_SegmentDateMin={date}&db_ShiftStart={shift}&db_ShiftEnd={shift}"
    # return "http://127.0.0.1:5500/assets/page-machine.html"

    if link_up == "17":
        params = {
            "table": "sp_PeriodEquipmentData",
            "act": "query",
            "eoa": "x",
            "db_Line": f"PMID-SE-CP-L0{link_up}",
            "db_SegmentDateMin": date,
            "db_ShiftStart": shift,
            "db_ShiftEnd": shift,
        }
    else:
        params = {
            "table": "sp_PeriodEquipmentData",
            "act": "query",
            "eoa": "x",
            "db_Line": f"ID01-SE-CP-L0{link_up}",
            "db_SegmentDateMin": date,
            "db_ShiftStart": shift,
            "db_ShiftEnd": shift,
        }

    url = MAIN_URL + "&".join(f"{key}={value}" for key, value in params.items())

    return url


def get_url_result(link_up: str, date: str, shift: str) -> str:
    """Faster and more efficient version of get_url_1"""
    # return f"http://ots.app.pmi/db.aspx?table=SPA_NormPeriodLossTree&act=query&db_Normalize=0&eoa=x&db_SegmentDateMin=2024-09-25&db_ShiftStart=2&db_ShiftEnd=2&db_Line=ID01-SE-CP-L027&db_Language=OEM"
    # return "http://127.0.0.1:5500/assets/page-shift.html"

    if link_up == "17":
        params = {
            "table": "SPA_NormPeriodLossTree",
            "act": "query",
            "db_Normalize": 0,
            "eoa": "x",
            "db_SegmentDateMin": date,
            "db_ShiftStart": shift,
            "db_ShiftEnd": shift,
            "db_Line": f"PMID-SE-CP-L0{link_up}",
            "db_Language": "OEM",
        }
    else:
        params = {
            "table": "SPA_NormPeriodLossTree",
            "act": "query",
            "db_Normalize": 0,
            "eoa": "x",
            "db_SegmentDateMin": date,
            "db_ShiftStart": shift,
            "db_ShiftEnd": shift,
            "db_Line": f"ID01-SE-CP-L0{link_up}",
            "db_Language": "OEM",
        }

    return MAIN_URL + "&".join(f"{k}={v}" for k, v in params.items())


def resource_path(relative_path: str) -> str:
    """Get absolute path to resource, works for dev and for PyInstaller

    Args:
        relative_path: The relative path to the resource.

    Returns:
        str: The absolute path to the resource.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_script_folder() -> str:
    """Get absolute path to the script folder, works for dev and for PyInstaller

    When running as a PyInstaller bundle, sys.executable is the path to the
    executable. Otherwise, we get the path to the script that was executed.

    Returns:
        str: The absolute path to the script folder.
    """

    if getattr(sys, "frozen", False):
        script_path: str = os.path.dirname(sys.executable)
    else:
        script_path: str = os.path.dirname(
            os.path.abspath(sys.modules["__main__"].__file__)
        )
    return script_path


def create_excel_file() -> None:
    file_path = os.path.join(get_script_folder(), "DB.xlsx")

    sheets_list = ["Data", "Username", "Link", "DailyTarget"]

    if not os.path.exists(file_path):
        with open(file_path, "w") as file:
            wb: Workbook = Workbook()
            sh_names = wb.sheetnames
            wb.active.title = "Data"
            for sh in sh_names:
                if not sh in sheets_list:
                    wb.create_sheet(title="Username")
                    wb.create_sheet(title="Link")
                    wb.create_sheet(title="DailyTarget")

            ws = wb["DailyTarget"]
            ws["A1"] = ""
            ws["A2"] = "STOP"
            ws["A3"] = "MTBF"
            ws["A4"] = "PR"
            ws["A5"] = "NATR"
            ws["A6"] = "PDT"
            ws["A7"] = "UPDT"
            ws["B1"] = "TARGET"

            wb.save(file_path)
            wb.close()
    else:
        pass


def get_excel_filename() -> str:
    file_path = os.path.join(get_script_folder(), "DB.xlsx")
    if not os.path.exists(file_path):
        create_excel_file()
        return file_path
    else:
        return file_path


def get_data_from_excel(sheet_index: int):
    # create_excel_file()
    names = []
    wb = openpyxl.load_workbook(os.path.join(get_script_folder(), "DB.xlsx"))
    sheet = wb.worksheets[sheet_index]

    for i, row in enumerate(sheet):
        # if i == 0:
        #     continue
        name = row[0].value
        names.append(name)

    wb.save("DB.xlsx")
    wb.close()

    return names
