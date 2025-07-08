import asyncio

import httpx
import pandas as pd
import ttkbootstrap as ttk
from async_tkinter_loop import async_handler
from openpyxl import load_workbook
from openpyxl.styles import Font
from tabulate import tabulate
from ttkbootstrap.constants import *

from src.core.logic import fetch_data, get_data_spa, get_time_period, read_csv
from src.gui.qr import generate_qrcode
from src.gui.target_editor import EditableTableview
from src.gui.toast import create_toast
from src.utils.constants import HEADERS, NTLM_AUTH
from src.utils.csvhandle import get_targets_file_path, load_targets_df
from src.utils.helpers import (
    get_data_from_excel,
    get_excel_filename,
    get_url_result,
    get_url_stop,
    read_config,
    resource_path,
)

from .mainpage import MainScreen
from .sidebar import Sidebar


class View(ttk.Window):
    def __init__(self) -> None:
        super().__init__(themename="superhero")
        self.title("Daily Report")
        self.iconbitmap(resource_path("assets/c5_spa.ico"))
        self.excelDB = ttk.StringVar(value=get_excel_filename())
        self.config = read_config()

        # Initialize Sidebar
        self.sidebar = Sidebar(self)
        self._configure_sidebar()
        self.sidebar.pack(side=LEFT, fill=Y)

        # Add separator
        ttk.Separator(
            self,
            orient="vertical",
            style="success.Vertical.TSeparator",
        ).pack(side=LEFT, fill=Y, pady=10)

        # Initialize Main Screen
        self.mainscreen = MainScreen(self)
        self.mainscreen.pack(side=LEFT, fill=BOTH, expand=YES)

    def _configure_sidebar(self):
        """Configure sidebar buttons and dropdowns."""
        self.sidebar.lu.configure(
            values=sorted(self.config.get("DEFAULT", "link_up").split(","))
        )
        self.sidebar.lu.current(0)

        self.sidebar.btn_get_data.configure(command=self.get_data)
        self.sidebar.btn_target.configure(command=self.show_target_editor)
        self.sidebar.btn_qr.configure(command=self.create_qrcode_toplevel)
        self.sidebar.btn_result.configure(command=self.show_result)
        self.sidebar.btn_save.configure(command=self.save_excel)

    def _get_url(self, endpoint_type, link_up, date_entry, shift):
        """Helper method to generate URLs based on environment."""
        if self.config.get("DEFAULT", "environment") == "production":
            if endpoint_type == "stop":
                return get_url_stop(link_up, date_entry, shift)
            elif endpoint_type == "result":
                return get_url_result(link_up, date_entry, shift)
        else:
            return f"http://127.0.0.1:5500/assets/page-{endpoint_type}.html"

    @async_handler
    async def get_data(self):
        """Fetch and display data based on user input."""
        if not self.sidebar.select_shift.get():
            create_toast("Select shift first", WARNING)
            return

        link_up = self.sidebar.lu.get().lstrip("LU")
        date_entry = self.sidebar.dt.entry.get()
        shift = self.sidebar.select_shift.get().lstrip("Shift ")
        url = self._get_url("stop", link_up, date_entry, shift)

        try:
            self.mainscreen.progressbar.start()
            self.sidebar.btn_get_data.configure(state=DISABLED)

            async with httpx.AsyncClient(timeout=30) as client:
                response = await client.get(
                    url,
                    follow_redirects=True,
                    headers=HEADERS,
                    auth=NTLM_AUTH,
                )
            if response.status_code == 200:
                time_period = get_time_period(response)
                self.mainscreen.time_period.configure(text=time_period)

                # df = extract_dataframe(response)
                df = get_data_spa(response)

                self._populate_table(df)

                create_toast(f"App setting\n{url}", SUCCESS)
            else:
                create_toast(
                    f"Error Code {response.status_code}: {response.text}", DANGER
                )
        except httpx.HTTPError as e:
            create_toast(f"HTTP Error: {e}", DANGER)
        finally:
            self.mainscreen.progressbar.stop()
            self.sidebar.btn_get_data.configure(state=NORMAL)

    def _populate_table(self, df: pd.DataFrame):
        """Populate the table with data."""
        head = df.columns.to_list()
        data = df.values.tolist()

        self.mainscreen.table.columnconfigure(0, weight=1)
        self.mainscreen.table.columnconfigure(1, weight=1)
        self.mainscreen.table.build_table_data(head, data)
        self.mainscreen.table.reset_table()

    @async_handler
    async def show_result(self):
        """Fetch and display result data."""
        if not self.sidebar.select_shift.get():
            create_toast("Select shift first", WARNING)
            return

        self.mainscreen.progressbar.start()

        link_up = self.sidebar.lu.get().lstrip("LU")
        date_entry = self.sidebar.dt.entry.get()
        shift = self.sidebar.select_shift.get().lstrip("Shift ")
        url = self._get_url("result", link_up, date_entry, shift)

        excel_file = get_targets_file_path(link_up)

        try:
            async with httpx.AsyncClient() as client:
                fetch_task = fetch_data(url, client)
                read_task = read_csv(excel_file, shift=shift)
                http_result, excel_result = await asyncio.gather(fetch_task, read_task)

            self._display_result(http_result, excel_result)
        except Exception as e:
            create_toast(f"Error: {e}", DANGER)
        finally:
            self.mainscreen.progressbar.stop()

    def _display_result(self, http_result, excel_result):
        """Display the result data in the UI."""
        data = [excel_result, http_result[0]]
        txt = tabulate(
            pd.DataFrame(data).transpose().reset_index().values.tolist(),
            tablefmt="pretty",
            headers=[self.sidebar.lu.get(), "TARGET", "ACTUAL"],
            numalign="left",
            stralign="left",
        )

        value = "`" + txt.replace("\n", "`\n`") + "`\n\n"

        if "+-" in self.mainscreen.inp.get("1.0", "2.0"):
            self.mainscreen.inp.delete("1.0", "12.0")
        self.mainscreen.inp.insert("1.0", value)
        create_toast("Updated", SUCCESS)

        self.mainscreen.time_period.configure(text=http_result[1])

    @async_handler
    async def create_qrcode_toplevel(self):
        """Create a QR code and display it in a new window."""
        width, height = 500, 500
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)

        try:
            self.toplevel.destroy()
        except AttributeError:
            pass

        self.toplevel = ttk.Toplevel(self, position=(x, y))
        self.toplevel.title("QR Code")

        qr_label = ttk.Label(self.toplevel)
        qr_label.pack(expand=True, fill=BOTH)

        tanggal = self.sidebar.dt.entry.get()
        shift = self.sidebar.select_shift.get()
        report = self.mainscreen.inp.get("1.0", ttk.END)

        header = f"{tanggal}, {shift}"

        text = f"{header}\n{report}"

        await generate_qrcode(text, qr_label)

    def show_target_editor(self):
        """Open the target editor window."""
        try:
            self.target.destroy()
        except AttributeError:
            pass

        self.target = ttk.Toplevel(self)
        self.target.title("Target Editor")

        lu = self.sidebar.lu.get().lstrip("LU")
        file_path = get_targets_file_path(lu)
        target_df = load_targets_df(file_path)

        columns = target_df.columns.to_list()
        data = target_df.values.tolist()

        table = EditableTableview(self.target, columns, data)
        table.pack(fill=BOTH, expand=False, padx=10, pady=10)

        save_btn = ttk.Button(
            self.target,
            text="Save",
            command=lambda: table.save_to_csv(file_path),
            bootstyle=SUCCESS,
        )
        save_btn.pack(pady=5)

    def save_excel(self):
        """Save data to the Excel file."""
        try:
            if not self.sidebar.entry_user.get():
                create_toast("Enter username first", WARNING)
                return

            file_excel = self.excelDB.get()
            wb = load_workbook(file_excel)

            sheet_data = wb["Data"]
            ID = len(sheet_data["No"])
            data = [
                ID,
                self.sidebar.dt.entry.get(),
                self.sidebar.select_shift.get(),
                self.sidebar.lu.get(),
                self.sidebar.entry_user.get(),
                self.mainscreen.inp.get("1.0", ttk.END),
            ]
            sheet_data.append(data)
            sheet_data.cell(row=ID + 1, column=6).font = Font(name="Consolas", size=10)

            sheet_setting = wb["Username"]
            list_username = get_data_from_excel(sheet_index=1)
            if self.sidebar.entry_user.get() not in list_username:
                sheet_setting.append([self.sidebar.entry_user.get()])
                list_username.append(self.sidebar.entry_user.get())
                self.sidebar.entry_user.configure(completevalues=list_username)

            wb.save(file_excel)
            create_toast("File is successfully updated.", SUCCESS)
        except PermissionError:
            create_toast(
                "File is being used by another User.\nPlease try again later.", DANGER
            )
